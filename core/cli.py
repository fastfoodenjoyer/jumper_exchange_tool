import os
import asyncio
import json
from json import JSONDecodeError
from pathlib import Path
from datetime import datetime

import questionary
from rich.console import Console
from rich.table import Table
from sqlalchemy.exc import OperationalError
from termcolor import cprint
from art import text2art
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.db_utils.db import db
from core.db_utils.models import RouteStatus, Route
from utils.utils import read_toml

console = Console()


def load_presets():
    presets = []
    presets_dir = Path("presets")
    for preset_file in presets_dir.glob("*.toml"):
        preset_data = read_toml(preset_file)
        params_str = json.dumps(preset_data["functions_params"],
                                separators=(', ', ':')).replace(' ', '\n').replace('[', '\n[')

        preset_json = {
            "name": preset_data["preset"]["name"],
            "path": preset_file,
            "description": preset_data["preset"]["description"],
            "params": params_str,
            "repeat_actions": preset_data.get("repeat_actions")
        }

        presets.append(preset_json)
    return presets


def display_presets_table(presets):
    table = Table(title="Available Presets", show_lines=True)
    table.add_column("Name", style="cyan")
    table.add_column("Description", style="green")
    table.add_column("Parameters", style="yellow")

    for preset in presets:
        table.add_row(preset["name"], preset["description"], preset["params"])

    console.print(table)


def display_accounts(route_status: RouteStatus):
    table = Table(title=f"{route_status.name} Accounts", show_lines=True)
    table.add_column("Name", style="cyan")
    table.add_column("Actions", style="green")
    table.add_column("Completed At Time", style="yellow")

    routes: list[Route] = db.get_routes_by_statuses([route_status])

    for route in routes:
        account = db.get_account_by_id(route.account_id)
        str_actions = [action.action_name for action in route.actions]
        str_actions = "Route Actions: " + ", ".join(str_actions)
        completed_at = route.completed_at.strftime("%d/%m/%Y, %H:%M:%S") if route.completed_at else "Not completed yet"

        table.add_row(account.name, str_actions, completed_at)

    console.print(table)

async def edit_preset_params():
    params_object = db.get_action_params()
    params = json.loads(params_object.action_params)
    console.print(f"Current parameters: {params}")
    new_params = await questionary.text(f"Enter edited parameters:").ask_async()
    new_params = new_params.strip().replace("True", "true").replace("False", "false").replace("None", "null").replace("'", '"')
    try:
        json.loads(new_params)
    except JSONDecodeError:
        console.print(f"[red]Invalid JSON format. Please enter valid JSON.[/red]")
        return

    await db.update_obj_column(params_object, "action_params", new_params)
    console.print(f"[green]Parameters updated successfully.[/green]")

async def display_accounts_paginated(route_status: RouteStatus):
    """
    Функция для постраничного отображения аккаунтов с возможностью навигации между страницами.
    """
    page = 1
    page_size = 50
    total_pages = 1

    while True:
        try:
            # Получаем количество страниц
            total_count = db.get_routes_count_by_statuses([route_status])
            total_pages = (total_count + page_size - 1) // page_size  # округление вверх

            # Если нет данных, показываем сообщение и возвращаемся в основное меню
            if total_count == 0:
                console.print(f"No accounts found with status {route_status.name}")
                await asyncio.sleep(2)
                return

            # Получаем данные для текущей страницы
            offset = (page - 1) * page_size
            routes = db.get_routes_by_statuses_paginated([route_status], page_size, offset)

            # Отображаем таблицу
            table = Table(title=f"{route_status.name} Accounts (Page {page} of {total_pages})", show_lines=True)
            table.add_column("Name", style="cyan")
            table.add_column("Actions", style="green")
            table.add_column("Completed At Time", style="yellow")

            for route in routes:
                grouped_actions = {}
                for action in route.actions:
                    if action.action_name not in grouped_actions:
                        grouped_actions[action.action_name] = 0
                    else:
                        grouped_actions[action.action_name] += 1

                names_list = []
                for action_name, count in grouped_actions.items():
                    if count > 1:
                        action_name += f" (x{count})"
                        names_list.append(action_name)

                str_actions = "Route Actions: " + ", ".join(names_list)
                account = db.get_account_by_id(route.account_id)
                completed_at = route.completed_at.strftime(
                    "%d/%m/%Y, %H:%M:%S") if route.completed_at else "Not completed yet"

                table.add_row(account.name, str_actions, completed_at)

            console.print(table)
            console.print(
                f"Page {page} of {total_pages}, showing records {offset + 1}-{min(offset + len(routes), total_count)} of {total_count}")

            # Варианты навигации
            choices = []
            if page > 1:
                choices.append("Previous page")
            if page < total_pages:
                choices.append("Next page")
            choices.append("Return to main menu")

            nav_choice = await questionary.select(
                "Navigation options:",
                choices=choices
            ).ask_async()

            if nav_choice == "Previous page":
                page -= 1
            elif nav_choice == "Next page":
                page += 1
            elif nav_choice == "Return to main menu":
                return

        except Exception as e:
            console.print(f"[red]Error displaying accounts: {str(e)}[/red]")
            await asyncio.sleep(2)
            return


async def export_accounts_results_to_excel():
    """
    Экспортирует данные аккаунтов и их действий в Excel файл.
    Столбцы: имя аккаунта, действие, статус, время завершения
    """
    try:
        # Создаем директорию для отчетов, если она не существует
        reports_dir = os.path.join(os.getcwd(), "reports")
        os.makedirs(reports_dir, exist_ok=True)

        # Создаем имя файла с текущей датой и временем
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = os.path.join(reports_dir, f"accounts_report_{current_time}.xlsx")

        # Создаем новую рабочую книгу и выбираем активный лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts Report"

        # Определяем стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        centered = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Добавляем заголовки
        headers = ["Account Name", "Action", "Status", "Completed At"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            cell.border = thin_border

        # Получаем все маршруты из базы данных
        all_routes = []
        for status in RouteStatus:
            routes = db.get_routes_by_statuses([status])
            all_routes.extend(routes)

        # Заполняем данные
        row_num = 2
        for route in all_routes:
            account = db.get_account_by_id(route.account_id)
            route_completed = route.completed_at.strftime("%d/%m/%Y, %H:%M:%S") if route.completed_at else "Not completed yet"

            # Если у маршрута нет действий, добавляем одну строку с информацией о маршруте
            if not route.actions:
                ws.cell(row=row_num, column=1, value=account.name).border = thin_border
                ws.cell(row=row_num, column=2, value="No actions").border = thin_border
                ws.cell(row=row_num, column=3, value=route.status.name).border = thin_border
                ws.cell(row=row_num, column=4, value=route_completed).border = thin_border
                row_num += 1
            else:
                # Для каждого действия добавляем отдельную строку
                for action in route.actions:
                    action_completed = action.completed_at.strftime(
                        "%d/%m/%Y, %H:%M:%S") if action.completed_at else "Not completed yet"

                    ws.cell(row=row_num, column=1, value=account.name).border = thin_border
                    ws.cell(row=row_num, column=2, value=action.action_name).border = thin_border
                    ws.cell(row=row_num, column=3, value=action.status.name).border = thin_border
                    ws.cell(row=row_num, column=4, value=action_completed).border = thin_border
                    row_num += 1

        # Автоматически регулируем ширину столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Сохраняем файл
        wb.save(filename)

        console.print(f"[green]Отчет успешно сохранен в файл: {filename}[/green]")
        return filename

    except Exception as e:
        console.print(f"[red]Ошибка при экспорте данных в Excel: {str(e)}[/red]")
        return None


async def main_menu():
    cprint(text2art(text="fastfoodsofts", font="crawford", chr_ignore=True, space=0),
           color='blue', attrs=['bold'])
    cprint("https://t.me/fastfoodsofts", color='green', on_color='on_grey', attrs=['bold'])
    print('\n')

    while True:
        try:
            pending_route_actions = db.get_routes_by_statuses([RouteStatus.PENDING, RouteStatus.IN_PROGRESS])[0].actions
            str_actions = [action.action_name if action.action_name else "No pending routes" for action in
                           pending_route_actions]
            str_actions = "Route Actions: " + ", ".join(set(str_actions))
        except (IndexError, OperationalError):
            str_actions = "No pending routes"

        choice = await questionary.select(
            "What would you like to do?",
            choices=[
                f"Continue previous route from database ({str_actions})",
                "Run Preset",
                "View Presets",
                "Rerun Failed Actions",
                "Edit Current Preset Parameters",
                "View All Accounts by Status",
                "Export Accounts Results to Excel",
                "Exit"
            ]
        ).ask_async()

        if "Continue previous route from database" in choice:
            return "Continue"

        if choice == "Run Preset":
            presets = load_presets()
            preset_names = [p["name"] for p in presets]
            preset_choice = await questionary.select(
                "Select a preset to run:",
                choices=preset_names + ["Back"]
            ).ask_async()

            if preset_choice != "Back":
                selected_preset = next(p for p in presets if p["name"] == preset_choice)
                continue_route_choice = await questionary.select(
                    # "Continue previous route from database or create new routes?",
                    "Create new routes",
                    choices=["Create new routes", "Back"]  # "Continue old routes",
                ).ask_async()
                if continue_route_choice != "Back":
                    if continue_route_choice == "Create new routes":
                        return selected_preset

        elif choice == "View Presets":
            display_presets_table(load_presets())

        elif choice == "Rerun Failed Actions":
            return choice

        elif choice == "Edit Current Preset Parameters":
            await edit_preset_params()

        elif choice == "View All Accounts by Status":
            status_choice = await questionary.select(
                "Select account status to view:",
                choices=[status.name for status in RouteStatus]
            ).ask_async()
            selected_status = RouteStatus[status_choice]
            await display_accounts_paginated(selected_status)

        elif choice == "Export Accounts Results to Excel":
            console.print("[yellow]Exporting accounts data to Excel file...[/yellow]")
            filename = await export_accounts_results_to_excel()
            if filename:
                open_file = await questionary.confirm(
                    f"Excel file created. Want to open it now?"
                ).ask_async()
                if open_file:
                    os.startfile(filename)

        elif choice == "Exit":
            return None
